import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import unicodedata
import chardet  # For automatic encoding detection

"""
This function is used to update the labels in specific columns of a CSV file based on a new label mapping provided in a separate file.
It processes one or more columns and replaces existing values with corresponding new labels from the mapping file.

Arguments:
- columns_to_update: A list of column names (strings) where the labels need to be updated.
- file_path: The path to the input CSV file that contains the data to be modified.
- new_label_file_path: The path to the CSV file that contains the new label mappings. This file must have a 'LABEL' column and a 'New Label' column.

The function reads the input data file, updates the specified columns using the mapping provided, 
and outputs a new CSV file with the modified data. The output file is saved with '_processed' appended to the original filename.

The new label mapping is provided in a separate CSV file, which must contain two columns:
- 'LABEL': The existing label in the data file.
- 'New Label': The new label that will replace the existing label.

Returns:
- A new CSV file with updated labels in the specified columns.
- Prints a log of each label update and the name of the file where the modified data is saved.

Contributor: Yuping Zheng 2024-10-30
last modified: 2025-1-19
"""

def normalize_dataframe(df, normalization_form='NFKC'):
    return df.applymap(lambda x: unicodedata.normalize(normalization_form, x) if isinstance(x, str) else x)

def read_csv_with_auto_encoding(file_path):
    """
    Reads a CSV file using automatically detected encoding with chardet.
    Applies Unicode normalization to the resulting DataFrame.
    """
    with open(file_path, 'rb') as f:
        raw_data = f.read()
    detected = chardet.detect(raw_data)
    encoding = detected['encoding']
    if not encoding:
        raise ValueError(f"Could not detect encoding for file: {file_path}")
    print(f"Detected encoding {encoding} for {file_path}")
    df = pd.read_csv(file_path, low_memory=False, encoding=encoding)
    df = normalize_dataframe(df)
    return df

def update_values_script(columns_to_update, file_path, new_label_file_path, partitioner='|'):
    """
    Updates specified columns in a CSV file with new labels provided in another CSV file.
    Modified rows are highlighted in the resulting Excel file, and changes are logged.
    """
    # === Setup ===
    # Resolve file paths for compatibility with relative paths
    file_path = os.path.abspath(file_path)
    new_label_file_path = os.path.abspath(new_label_file_path)

    # Validate the partitioner input
    if not isinstance(partitioner, str) or len(partitioner) == 0:
        raise ValueError("Partitioner must be a non-empty string.")

    # Check file existence before proceeding
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    if not os.path.exists(new_label_file_path):
        raise FileNotFoundError(f"The file '{new_label_file_path}' does not exist.")

    # === Load Data ===
    # Use automatic encoding detection to read CSV files
    df = read_csv_with_auto_encoding(file_path)
    new_labels_df = read_csv_with_auto_encoding(new_label_file_path)

    # === Validate Input ===
    # Check if new labels file has required columns
    if 'LABEL' not in new_labels_df.columns or 'New Label' not in new_labels_df.columns:
        raise ValueError("The new labels file must contain 'LABEL' and 'New Label' columns.")

    # Create a mapping dictionary from old labels to new labels
    new_labels_dict = dict(zip(new_labels_df['LABEL'], new_labels_df['New Label']))

    # Ensure the specified columns exist in the input data
    missing_columns = [col for col in columns_to_update if col not in df.columns]
    if missing_columns:
        raise ValueError(f"The following columns are missing in the input data: {', '.join(missing_columns)}")

    # === Initialize Logging ===
    # Prepare a log file to track modifications
    log_file_path = os.path.splitext(file_path)[0] + "_modifications.txt"
    modifications_log = []

    def label_update(cell_value, column_name, row_idx, col_idx, modified_flags):
        """
        Updates cell values based on a dictionary of new labels.
        Parameters:
        - cell_value: Original value of the cell.
        - column_name: Name of the column.
        - row_idx: Row index of the cell.
        - col_idx: Column index of the cell.
        - modified_flags: Tracks whether a row was modified.
        """
        # === Handle Edge Cases ===
        # Skip null or NaN values
        if pd.isnull(cell_value):
            return cell_value

        # Ensure cell value is a string and strip whitespace
        cell_value = str(cell_value).strip()

        # Handle empty strings or 'nan' (case-insensitive)
        if cell_value.lower() in ('', 'nan'):
            return cell_value

        # === Tokenize and Update ===
        # Split the cell value into tokens using the specified partitioner
        tokens = cell_value.split(partitioner)
        updated_tokens = []
        modified = False

        # Replace tokens using the new labels dictionary
        for token in tokens:
            updated_token = token
            for key, value in new_labels_dict.items():
                if key in token:
                    updated_token = token.replace(key, value)
                    if not modified:
                        # Log the modification for the first update in this cell
                        modifications_log.append(
                            f"Row: {row_idx + 1}, Column: '{column_name}', Original: '{cell_value}', Modified: '{updated_token}'"
                        )
                    modified = True
            updated_tokens.append(updated_token)

        # Mark the row as modified if changes occurred
        if modified:
            modified_flags[row_idx] = 'Modified'

        # Join the updated tokens using the specified partitioner
        return partitioner.join(updated_tokens)

    # === Apply Updates ===
    # Track which rows have been modified
    modified_flags = ['Unchanged'] * len(df)

    # Update each specified column in the DataFrame
    for col_idx, column in enumerate(columns_to_update):
        if column in df.columns:
            df[column] = df.apply(lambda row: label_update(row[column], column, row.name, col_idx, modified_flags), axis=1)

    # Add a "Modified" column to the DataFrame for tracking
    df['Modified'] = modified_flags

    # === Save Results ===
    # Save the updated data to a new CSV file with UTF-8 encoding
    dir_name, base_name = os.path.split(file_path)
    name, ext = os.path.splitext(base_name)
    output_file = os.path.join(dir_name, f"{name}_processed{ext}")
    df.to_csv(output_file, index=False, encoding='utf-8')
    print(f"Modified data saved to {output_file}")

    # Save the modification log to a text file
    if modifications_log:
        with open(log_file_path, 'w', encoding='utf-8') as log_file:
            log_file.write("\n".join(modifications_log))
        print(f"Modifications log saved to {log_file_path}")
    else:
        print("No modifications were made; no log file created.")

    # === Highlight Modifications (Excel) ===
    # Optionally, highlight modified rows in an Excel file
    try:
        # Save the updated data as an Excel file
        excel_file = os.path.splitext(output_file)[0] + ".xlsx"
        df.to_excel(excel_file, index=False, engine='openpyxl')
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Define a yellow fill style for highlighting modified rows
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_idx, status in enumerate(modified_flags, start=2):  # Skip the header row
            if status == 'Modified':
                for col_idx in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = fill

        # Save the highlighted Excel file
        workbook.save(excel_file)
        print(f"Excel file with highlighting saved to {excel_file}")
    except Exception as e:
        print(f"Failed to apply Excel highlighting: {e}")

# === Example Usage ===
if __name__ == "__main__":
    # Example Usage:

    # 1. Define the path to your input CSV file (the data file you want to modify).
    #    Replace 'your_input_file.csv' with the actual file path to your data file.
    file_path = 'your_input_file.csv'

    # 2. Define the path to the CSV file containing the new label mappings.
    #    Replace 'your_new_label_file.csv' with the actual file path to your label mapping file.
    new_label_file_path = 'your_new_label_file.csv'

    # 3. Specify the list of column names where label updates are required.
    #    Modify the list below with the column names you want to process.
    
    columns_to_update = ['column_1', 'column_2', 'column_3']

    # 4. Specify the partitioner to use for tokenization (default is '|')
    partitioner = '|'

    # 5. Call the `update_values_script` function with the specified arguments.
    #    This will update the specified columns in the input file based on the new label mappings,
    #    generate a log file for modifications, and save the modified data.
    update_values_script(columns_to_update, file_path, new_label_file_path, partitioner)

'''# Example:
# Define file paths for the input data and new labels
file_path = r"C:\Users\00000\VO1\src\templates\individuals.csv"
new_label_file_path = r'C:\Users\00000\VO1\experimental\data\data_laurel\new_old.csv'

# Specify the columns to update
columns_to_update = ['alternative label (zh)']

# Specify the partitioner to use for tokenization (default is '|')
partitioner = '|'

# Run the script
update_values_script(columns_to_update, file_path, new_label_file_path, partitioner)
'''
